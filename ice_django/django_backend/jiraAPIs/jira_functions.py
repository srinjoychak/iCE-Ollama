import pymongo
import base64
import requests
import json
import os
import re
import requests
import xlsxwriter
import json
import pandas as pd
from atlassian import Jira
from datetime import datetime, timezone
from urllib.parse import urlparse

# from .jira_settings import JQL_QUERY, JIRA_ENDPOINTS
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from requests.auth import HTTPBasicAuth
from django_backend.settings import mongo_conn
import logging
from logging.config import dictConfig

from django_backend.settings import LOGGING

dictConfig(LOGGING)
logger = logging.getLogger("dev_logger")
db = mongo_conn()
JIRA_ENDPOINTS = {
    "search_endpoint": "{0}/rest/api/2/search",
}
JQL_QUERY = {
    "jql": "project = {0} AND sprint = {1}",
    "daily_scrum_report_jql": 'project = "{0}" AND "Scrum Team" = "{1}" AND sprint = {2} AND status NOT IN (Done, Closed)',
    "startAt": 0,
    "maxResults": 15,
    "fields": [
        "id",
        "key",
        "assignee",
        "summary",
        "description",
        "updated",
        "worklog",
        "comment",
        "status",
        "issuetype",
        "customfield_10002",
        "priority",
        "customfield_10026",
    ],
    "scrum_fields": [
        "key",
        "assignee",
        "summary",
        "reporter",
        "description",
        "updated",
        "timespent",
        "customfield_10034",
        "status",
        "issuetype",
        "updated",
        "priority",
    ],
}


class AuthenticationError(Exception):
    pass


class AuthorizationError(Exception):
    pass


def decode_auth(encoded_token):
    decoded_token = base64.b64decode(encoded_token).decode()
    return decoded_token


class JiraInstances:
    def __init__(self, project_name, user_name):
        self.db = db
        self.project_name = project_name
        self.user_name = user_name

    def get_jira_instances(self):
        admin_data = self.db.ice_webex_jira_project_data.find(
            {"project_name": self.project_name},
            {
                "Scrum_master": 1,
                "jire_host_url": 1,
                "board_name": 1,
                "jira_type": 1,
                "Project_key": 1,
                "borad_id": 1,
                "_id": 0,
            },
        )
        for token in admin_data:
            jire_host_url = token["jire_host_url"]
            self.board_name = token["board_name"]
            self.project_key = token["Project_key"]
            self.jira_type = token["jira_type"]
            try:
                self.board_id = token["borad_id"]
            except:
                self.board_id = None
            if self.jira_type == "server":
                self.host_url = "https://{0}/jira".format(jire_host_url)
                project_data = self.db.ice_user_data.find(
                    {"user_name": self.user_name},
                    {
                        "iCE_cred_properties.sso_auth": 1,
                        "my_primary_projects": 1,
                        "_id": 0,
                    },
                )
                for j in project_data:
                    auth_data = j["iCE_cred_properties"]["sso_auth"]
                    auth_data = decode_auth(auth_data)
                    self.basicauth_1 = "Basic {}".format(auth_data)
                    base64_bytes = auth_data.encode("ascii")
                    sample_string_bytes = base64.b64decode(base64_bytes)
                    sample_string = sample_string_bytes.decode("ascii")
                    auth = sample_string.split(":")
                    self.username = auth[0]
                    self.password = auth[1]
                    self.jira = Jira(
                        url=self.host_url, username=auth[0], password=auth[1]
                    )
                return [
                    self.jira,
                    self.host_url,
                    self.basicauth_1,
                    self.board_name,
                    self.username,
                    self.password,
                    self.project_key,
                    self.jira_type,
                    self.board_id,
                ]
            else:
                self.host_url = "https://{0}".format(jire_host_url)
                check_token = db.ice_user_data.find(
                    {"user_name": self.user_name},
                    {"iCE_cred_properties.api_token": 1, "email": 1, "_id": 0},
                )
                for token in check_token:
                    if token["iCE_cred_properties"]["api_token"] != "":
                        check_project = db.ice_user_data.find(
                            {"user_name": self.user_name},
                            {"my_primary_projects": 1, "_id": 0},
                        )
                        username = token["email"]
                        for project in check_project:
                            if project["my_primary_projects"] == "":
                                return [
                                    None,
                                    "Issue with empty primary project name. Please Add primary project again.",
                                ]
                            else:
                                self.username = username
                                self.basicauth_1 = token["iCE_cred_properties"][
                                    "api_token"
                                ]
                                try:
                                    self.jira = Jira(
                                        url=self.host_url,
                                        username=self.username,
                                        password=self.basicauth_1,
                                    )
                                except:
                                    return [None, "Credenticals issue"]
                    else:
                        return [None, "Empty API Token Please update the API Token"]
                return [
                    self.jira,
                    self.host_url,
                    self.basicauth_1,
                    self.board_name,
                    self.username,
                    None,
                    self.project_key,
                    self.jira_type,
                    self.board_id,
                ]


class JiraUpdater(JiraInstances):
    def __init__(self, project_name, user_name):
        super().__init__(project_name, user_name)

    def update_worklogs(self, issuekey, worklogcomment, comment, worklogduration):
        jira_data = self.get_jira_instances()
        if jira_data[0] == None:
            return jira_data[1]
        else:
            try:
                host_url = jira_data[1]
                basicauth = jira_data[2]
                jira_type = jira_data[7]
                username = jira_data[4]
                jira = jira_data[0]
                if jira_type == "server":
                    url = "{1}/rest/api/2/issue/{0}/worklog".format(issuekey, host_url)
                    payload = json.dumps(
                        {"comment": worklogcomment, "timeSpent": worklogduration}
                    )
                    headers = {
                        "Accept": "application/json",
                        "Content-Type": "application/json",
                        "Authorization": basicauth,
                    }
                    requests.request("POST", url, headers=headers, data=payload)
                else:
                    url = "{1}/rest/api/2/issue/{0}/worklog".format(issuekey, host_url)
                    print(url)
                    auth = HTTPBasicAuth(
                        username,
                        basicauth,
                    )
                    headers = {"Content-Type": "application/json"}
                    payload = {
                        "timeSpentSeconds": worklogduration,
                        "comment": {
                            "type": "doc",
                            "version": 1,
                            "content": [
                                {
                                    "type": "paragraph",
                                    "content": [
                                        {"type": "text", "text": worklogcomment}
                                    ],
                                }
                            ],
                        },
                    }
                    jira.issue_add_comment(issuekey, comment)
                    response = requests.post(
                        url, json=payload, auth=auth, headers=headers
                    )
                    logger.info(response)
                    if response.status_code == 200:
                        return "Issue Updated successfully"
                    elif response.status_code == 201:
                        return "Issue Updated successfully"
                    elif response.status_code == 401:
                        return "Invalid username or password."
                    elif response.status_code == 403:
                        return "You do not have permission to access the requested resource."
                    elif response.status_code == 400:
                        return "Bad Request Payload Data"
            except Exception as e:
                return "Error while adding worklog"


class JiraProjectRegistration:
    def __init__(self):
        self.db = db

    def validating_jira_instance(self, input_data):
        if input_data.startswith("https://"):
            input_data = input_data[len("https://") :]
        # Remove any leading slashes
        input_data = input_data.lstrip("/")
        # Remove any trailing slashes and anything that follows
        input_data = input_data.split("/", 1)[0]
        return input_data

    def extract_jira_info(self, url):
        # This regular expression will match the domain name and the board number in the Jira URL
        match = re.search(
            r"atlassian\.net/jira/software/(?:c/)?projects/(\w+)/boards/(\d+)", url
        )
        if match:
            project_key = match.group(1)
            board_number = match.group(2)
            parsed_url = urlparse(url)
            domain_name = parsed_url.netloc
            return [project_key, domain_name, board_number]
        else:
            return None

    def verify_project_key(
        self,
        user_name,
        project_key,
        host_url,
        board_name,
        github_url,
        scrum_master,
        server_type,
        board_id,
    ):
        project_data = self.db.ice_webex_jira_project_data.find(
            {}, {"Project_key": 1, "_id": 0}
        )
        project_list = []
        for i in project_data:
            project_list.append(i["Project_key"])
        if project_key in project_list:
            return "Project Already Exists"
        else:
            try:
                check_token = self.db.ice_user_data.find(
                    {"user_name": user_name},
                    {"iCE_cred_properties.sso_auth": 1, "_id": 0},
                )
                for token in check_token:
                    if token["iCE_cred_properties"]["sso_auth"] != "":
                        decode_auth_data = token["iCE_cred_properties"]["sso_auth"]
                        auth_data = base64.b64decode(decode_auth_data).decode()
                        base64_bytes = auth_data.encode("ascii")
                        sample_string_bytes = base64.b64decode(base64_bytes)
                        sample_string = sample_string_bytes.decode("ascii")
                        auth_data = sample_string.split(":")
                        user_name = auth_data[0]
                        password = auth_data[1]
                        jire_host_url = "https://{0}/jira".format(host_url)
                        jira = Jira(
                            url=jire_host_url, username=user_name, password=password
                        )
                        response = jira.get_all_projects()
                        for i in range(len(response)):
                            if response[i]["key"] == project_key:
                                project_name = response[i]["name"]
                            else:
                                return "Enter valid Project Key"
                        self.db.ice_webex_jira_project_data.insert_one(
                            {
                                "project_name": project_name,
                                "Project_key": project_key,
                                "jire_host_url": jire_host_url,
                                "board_name": board_name,
                                "borad_id": "",
                                "jira_type": server_type,
                                "scanflag": "True",
                                "schedulerflag": "True",
                                "git_repo_url": github_url,
                                "active_sprint_info": [
                                    {
                                        "active_gitbranch_name": "",
                                        "Sonar_result": "[]",
                                        "Coronascan_results": "[]",
                                        "active_sprint_name": "",
                                        "sprint_id": "",
                                        "startdate": "",
                                        "enddate": "",
                                        "sprint_status": "",
                                    }
                                ],
                                "jenkins_config": [],
                                "selected_somaf_url": "",
                                "somaf_details": [],
                                "csdl_details": [],
                                "generic_credentials": [],
                                "Scrum_master": [{"username": scrum_master}],
                                "schudular_time_notification_time": "",
                                "projectcreated": datetime.today().strftime(
                                    "%d-%m-%Y %I:%M %p"
                                ),
                                "projectlastActivity": datetime.today().strftime(
                                    "%d-%m-%Y %I:%M %p"
                                ),
                            }
                        )
                        return "Project Registration Successful"
                    else:
                        return "Add SSO Auth Token"
            except Exception as e:
                return "Enter Valid Parameters."

    def verify_board_name(
        self,
        user_name,
        board_name,
        host_url,
        project_key,
        github_url,
        scrum_master,
        server_type,
        board_id,
    ):
        try:
            project_data = self.db.ice_webex_jira_project_data.find(
                {}, {"board_name": 1, "_id": 0}
            )
            project_board_list = []
            logger.info(project_data)
            for i in project_data:
                project_board_list.append(i["board_name"])
            logger.info(project_board_list)
            if board_name in project_board_list:
                return [True, "Board Name Already Exists"]
            else:
                check_token = self.db.ice_user_data.find(
                    {"user_name": user_name},
                    {"iCE_cred_properties.api_token": 1, "email": 1, "_id": 0},
                )
                for token in check_token:
                    if token["iCE_cred_properties"]["api_token"] != "":
                        token_api = token["iCE_cred_properties"]["api_token"]
                        email = token["email"]
                        host_url = self.validating_jira_instance(host_url)
                        logger.info(host_url)
                        jire_host_url = "https://{0}".format(host_url)
                        jira = Jira(
                            url=jire_host_url, username=email, password=token_api
                        )
                        response = jira.get_all_projects()
                        project_list = []
                        for i in range(len(response)):
                            project_list.append(response[i]["key"])
                        if project_key in project_list:
                            project_name = response[i]["name"]
                        else:
                            return [False, "Enter valid Project Key"]
                        self.db.ice_webex_jira_project_data.insert_one(
                            {
                                "project_name": board_name,
                                "Project_key": project_key,
                                "jire_host_url": host_url,
                                "board_name": board_name,
                                "jira_type": server_type,
                                "borad_id": board_id,
                                "scanflag": "True",
                                "schedulerflag": "True",
                                "git_repo_url": github_url,
                                "active_sprint_info": [
                                    {
                                        "active_gitbranch_name": "",
                                        "Sonar_result": "[]",
                                        "Coronascan_results": "[]",
                                        "active_sprint_name": "",
                                        "sprint_id": "",
                                        "startdate": "",
                                        "enddate": "",
                                        "sprint_status": "",
                                    }
                                ],
                                "jenkins_config": [],
                                "somaf_details": [],
                                "csdl_details": [],
                                "generic_credentials": [],
                                "Scrum_master": [{"username": scrum_master}],
                                "schudular_time_notification_time": "",
                                "projectcreated": datetime.today().strftime(
                                    "%d-%m-%Y %I:%M %p"
                                ),
                                "projectlastActivity": datetime.today().strftime(
                                    "%d-%m-%Y %I:%M %p"
                                ),
                            }
                        )
                        return [True, "Project Registration Successful"]
                    else:
                        return [False, "Update API Token"]
        except Exception as e:
            return [False, "Enter Valid Parameters."]


class JiraQuery(JiraInstances):
    
    def __init__(self, project_name, user_name):
        super().__init__(project_name, user_name)

    def run_jql_query(self, url, jql, fields, max_result, jira_type):
        try:
            jira_data = self.get_jira_instances()
            if jira_data[0] == None:
                return jira_data[1]
            else:
                base_auth = jira_data[2]
                username = jira_data[4]
                if jira_type == "server":
                    query = json.dumps(
                        {
                            "jql": jql,
                            "startAt": 0,
                            "maxResults": max_result,
                            "fields": fields,
                        }
                    )
                    headers = {
                        "Authorization": f"{base_auth}",
                        "Content-Type": "application/json",
                    }

                    response = requests.post(url, headers=headers, data=query)
                    data = response.json()["issues"]
                    return data
                else:
                    headers = {
                        "Accept": "application/json",
                        "Content-Type": "application/json",
                    }
                    query_params = {
                        "jql": jql,
                        "startAt": 0,
                        "fields": fields,
                        "maxResults": max_result,
                    }
                    response = requests.get(
                        url,
                        headers=headers,
                        auth=(username, base_auth),
                        params=query_params,
                    )
                    if response.status_code == 401:
                        # Handle invalid credentials
                        raise AuthenticationError("Invalid username or password.")
                    elif response.status_code == 403:
                        # Handle insufficient permissions
                        raise AuthorizationError(
                            "You do not have permission to access the requested resource."
                        )
                    elif response.status_code == 200:
                        data = response.json()["issues"]
                        return data
                    else:
                        return "Getting Error when run the JQl Query Failed"
        except Exception as e:
            return "Getting Error when run the JQl Query Failed"
        except AuthenticationError as e:
            return "Invalid username or password."
        except AuthorizationError as e:
            return "You do not have permission to access the requested resource."

    def get_active_sprint_info(self):
        jira_data = self.get_jira_instances()
        if jira_data[0] == None:
            return jira_data[1]
        else:
            project_key = jira_data[6]
            jira_type = jira_data[7]
            jira_host = jira_data[1]
            basic_auth = jira_data[2]
            board_anme = jira_data[3]
            user_name = jira_data[4]
            board_id = jira_data[8]
            sprint_data = get_active_sprint_issues(
                jira_host, basic_auth, board_anme, jira_type, user_name, board_id
            )
            if sprint_data[0] != None:
                jql_query = JQL_QUERY.get("jql").format(project_key, sprint_data[2])
                # Splitting the input string to extract project and sprint values
                parts = jql_query.split("AND")
                project_part = parts[0].strip()
                sprint_part = parts[1].strip()
                # Adding single quotes around the sprint value
                sprint_value = "'" + sprint_part.split("=")[1].strip() + "'"
                # Creating the output string
                jql_query = f"{project_part} AND sprint = {sprint_value}"
                fields = JQL_QUERY.get("fields")
                max_results = JQL_QUERY.get("max_results")
                host_url = jira_data[1]
                url = JIRA_ENDPOINTS.get("search_endpoint").format(host_url)
                return self.run_jql_query(
                    url, jql_query, fields, max_results, jira_type
                )
            else:
                return sprint_data[1]

    def get_scrum_report_data(self):
        jira_data = self.get_jira_instances()
        if jira_data[0] == None:
            return jira_data[1]
        else:
            project_key = jira_data[6]
            jira_type = jira_data[7]
            jira_host = jira_data[1]
            basic_auth = jira_data[2]
            board_name = jira_data[3]
            user_name = jira_data[4]
            board_id = jira_data[8]
            sprint_data = get_active_sprint_issues(
                jira_host, basic_auth, board_name, jira_type, user_name, board_id
            )
            if sprint_data[0] != None:
                jql_query = JQL_QUERY.get("daily_scrum_report_jql").format(
                    project_key, board_name, sprint_data[0]
                )
                fields = JQL_QUERY.get("scrum_fields")
                max_results = JQL_QUERY.get("max_results")
                host_url = jira_data[1]
                url = JIRA_ENDPOINTS.get("search_endpoint").format(host_url)
                data = self.run_jql_query(
                    url, jql_query, fields, max_results, jira_type
                )
                logger.info(f"data: {data}")
                issue_data = []
                for issue in data:
                    issue_key = issue["key"]
                    text1 = "{1}/browse/{0}".format(issue_key, jira_host)
                    id_link = '=HYPERLINK("{0}", "{1}")'.format(text1, issue_key)
                    summary = issue["fields"]["summary"]
                    status = issue["fields"]["status"]["name"]
                    description = issue["fields"]["description"]
                    priority = issue["fields"]["priority"]["name"]
                    issuetype = issue["fields"]["issuetype"]["name"]
                    assignee = issue["fields"]["assignee"]["displayName"]
                    reporter = issue["fields"]["reporter"]["displayName"]
                    comments = issue["fields"].get("customfield_10034")
                    if comments == None:
                        comments = "No comments Found"
                    timespent = issue["fields"].get("timespent")
                    if timespent == None:
                        timespent = "Not Added any worklog"
                    else:
                        timespent = convert_seconds_to_time(timespent)
                    last_updates = issue["fields"]["updated"]
                    if last_updates == None:
                        last_updates = "No updated date"
                    else:
                        last_updates = time_conversion(last_updates)
                    issue_data.append(
                        {
                            "Issue Key": id_link,
                            "keys": issue_key,
                            "Issue Type": issuetype,
                            "Status": status,
                            "Description": description,
                            "Comments": comments,
                            "Assignee": assignee,
                            "Reporter": reporter,
                            "Total Worklog Effort": timespent,
                            "last_updated_date": last_updates,
                            "Priority": priority,
                        }
                    )
                return issue_data
            else:
                return sprint_data[1]


def extract_time(timeSpent):
    parts = timeSpent.split()
    days, hours, minutes = 0, 0, 0
    for part in parts:
        if "d" in part:
            days = int(part[:-1])
        elif "h" in part:
            hours = int(part[:-1])
        elif "m" in part:
            minutes = int(part[:-1])
    return [days, hours, minutes]



def get_total_time(time):
    time_spent = []
    for d in time:
        time_spent.append(d["timeSpent"])
    time_spent = list(filter(None, time_spent))
    if len(time_spent) == 0:
        return int("0")
    else:
        time_list = [extract_time(i) for i in time_spent]
        total_minutes = sum(
            minutes + hours * 60
            for days, hours, minutes in time_list
            for hours, minutes in [(hours + days * 8, minutes)]
        )
        total_hours = total_minutes // 60
        total_minutes = total_minutes % 60
        return f"{total_hours}h {total_minutes}m"


def worklog_func(worklogs):
    worklog = []
    for d in worklogs:
        try:
            worklog.append(d["comment"])
        except:
            worklog = []
    worklog = list(filter(None, worklog))
    if len(worklog) == 0:
        return "No comments found"

    else:
        return worklog[-1]


def comment_func(comments):
    comment = []
    for d in comments:
        try:
            comment.append(d["body"])
        except:
            comment = []
    comment = list(filter(None, comment))
    if len(comment) == 0:
        return "No comments found"

    else:
        return comment[-1]


def sprint_json_data(df):
    required_fields = {
        "key",
        "fields_issuetype_name",
        "fields_status_name",
        "fields_summary",
        "fields_comment_comments",
        "fields_description",
        "fields_assignee_emailAddress",
        "fields_updated",
        "fields_priority_name",
        "fields_worklog_worklogs",
    }
    diff_field = required_fields.difference(df.columns.to_list())
    for field in diff_field:
        df[field] = ""
    df["Worklog"] = df["fields_worklog_worklogs"].apply(
        lambda x: worklog_func(x) if len(x) != 0 else "No comments found"
    )
    df["comments"] = df["fields_comment_comments"].apply(
        lambda x: comment_func(x) if len(x) != 0 else "No comments found"
    )
    df["Total Worklog Effort(hours)"] = df["fields_worklog_worklogs"].apply(
        lambda x: get_total_time(x) if len(x) != 0 else int("0")
    )
    try:
        df["fields_updated"] = pd.to_datetime(
            df["fields_updated"], format=("%Y-%m-%dT%H:%M:%S.%f")
        )
        df["fields_updated"] = df["fields_updated"].dt.tz_localize(None)
    except Exception as e:
        pass
    sprint_df = df.rename(
        columns={
            "fields_issuetype_name": "Issue Type",
            "key": "Key",
            "fields_status_name": "Status",
            "fields_summary": "Summary",
            "comments": "Comment",
            "fields_description": "Description",
            "fields_assignee_emailAddress": "Assignee",
            "Worklog": "Worklog",
            "Total Worklog Effort(hours)": "Total Worklog Effort (hours)",
            "fields_updated": "Updated",
        }
    )[
        [
            "Issue Type",
            "Key",
            "Status",
            "Summary",
            "Comment",
            "Description",
            "Assignee",
            "Worklog",
            "Total Worklog Effort (hours)",
            "Updated",
        ]
    ]
    data = sprint_df.to_json(orient="records")
    return data


def get_active_sprint_issues(
    host_url, basicauth_1, boardName, jira_type, username, board_id
):
    try:
        if jira_type == "server":
            board_id = server_get_board_id(host_url, basicauth_1, boardName)
            api_url = "{1}/rest/agile/1.0/board/{0}/sprint?state=active".format(
                board_id[0], host_url
            )
            headers = {"Authorization": basicauth_1}
            # Make API call to get active sprint information
            response = requests.get(api_url, headers=headers)
            if response.status_code == 200:
                sprint_data = response.json()
                sprint_name = sprint_data["values"][0]["name"]
                startDate = sprint_data["values"][0]["startDate"]
                from datetime import datetime

                datetime_object = datetime.strptime(startDate, "%Y-%m-%dT%H:%M:%S.%fZ")
                startDate = datetime_object.strftime("%Y-%m-%d")
                endDate = sprint_data["values"][0]["endDate"]
                datetime_object = datetime.strptime(endDate, "%Y-%m-%dT%H:%M:%S.%fZ")
                endDate = datetime_object.strftime("%Y-%m-%d")
                # Extract Sprint ID from response
                sprint_id = sprint_data["values"][0]["id"]
                return [sprint_id, board_id[0], sprint_name, startDate, endDate]
        else:
            # board_id = cloud_get_board_id(host_url, basicauth_1, boardName, username)
            headers = {"Accept": "application/json", "Content-Type": "application/json"}

            # Construct the API endpoint URL
            endpoint_url = "{1}/rest/agile/1.0/board/{0}/sprint?state=active".format(
                board_id, host_url
            )
            # Make the GET request
            response = requests.get(
                endpoint_url, headers=headers, auth=(username, basicauth_1)
            )
        # Check if the request was successful
        if response.status_code == 401:
            # Handle invalid credentials
            raise AuthenticationError("Invalid username or password.")
        elif response.status_code == 403:
            # Handle insufficient permissions
            raise AuthorizationError(
                "You do not have permission to access the requested resource."
            )
        elif response.status_code == 200:
            try:
                sprint_data = response.json()
                sprint_name = sprint_data["values"][0]["name"]
                startDate = sprint_data["values"][0]["startDate"]
                from datetime import datetime

                datetime_object = datetime.strptime(startDate, "%Y-%m-%dT%H:%M:%S.%fZ")
                startDate = datetime_object.strftime("%Y-%m-%d")
                endDate = sprint_data["values"][0]["endDate"]
                datetime_object = datetime.strptime(endDate, "%Y-%m-%dT%H:%M:%S.%fZ")
                endDate = datetime_object.strftime("%Y-%m-%d")
                # Extract Sprint ID from response
                sprint_id = sprint_data["values"][0]["id"]
                return [sprint_id, board_id, sprint_name, startDate, endDate]
            except Exception as e:
                return [
                    None,
                    "Currently, there are no active sprints available.",
                ]
        else:
            return [
                None,
                "The board does not provide support for sprint-based project management frameworks.",
            ]
    except AuthenticationError as e:
        return [None, "Invalid username or password."]
    except AuthorizationError as e:
        return [None, "You do not have permission to access the requested resource."]
    except Exception as e:
        return [None, str(e)]


## jira_server boardid
def server_get_board_id(host_url, basicauth_1, board_name):
    url = "{0}/rest/agile/1.0/board?startAt=0".format(host_url)
    headers = {"Authorization": basicauth_1}
    response = requests.request("GET", url, headers=headers)
    if response.status_code == 200:
        data = response.json()
        num_of_itr = round(data["total"] / 50 + 1)
        boardiddict = {}
        for i in range(0, num_of_itr + 1):
            startAt = i * 50
            board_url = "{1}/rest/agile/1.0/board?startAt={0}".format(startAt, host_url)
            headers = {"Authorization": basicauth_1}
            response = requests.request("GET", board_url, headers=headers, data={})
            data = response.json()
            for i in range(0, len(data["values"])):
                boardiddict.update({data["values"][i]["name"]: data["values"][i]["id"]})
        board_id = boardiddict[board_name]
        return [board_id]
    else:
        return "invali data"


def cloud_get_board_id(board_id):
    return [board_id]


def add_project_to_primaryproject(project_name, user_name):

    project_data = db.ice_webex_jira_project_data.find(
        {"project_name": project_name}, {"_id": 1}
    )
    for i in project_data:
        id = i["_id"]
    db.ice_user_data.update_one(
        {"user_name": user_name},
        {"$push": {"my_jira_projects": {str(id): project_name}}},
    )
    data = db.ice_user_data.find({}, {"my_primary_projects": 1, "_id": 0})
    for j in data:
        my_primary_projects = j["my_primary_projects"]
    if project_name != my_primary_projects:
        db.ice_user_data.update_one(
            {"user_name": user_name},
            {"$set": {"my_primary_projects": project_name}},
        )


def check_ice_inventory(project_name, user_name):
    jira_data = db.ice_user_data.find(
        {"user_name": user_name}, {"my_jira_projects": 1, "_id": 0}
    )
    my_jira_project_list = []
    try:
        for i in jira_data:
            list_jira = i["my_jira_projects"]
        for j in list_jira:
            for k, v in j.items():
                my_jira_project_list.append(v)
        if project_name in my_jira_project_list:
            # add_project_to_primaryproject(project_name, sender_id)
            return "Project Already Exit in Your iCE inventory"
        else:
            add_project_to_primaryproject(project_name, user_name)
            return None
    except Exception as e:
        logger.info(e)
        add_project_to_primaryproject(project_name, user_name)
        return None


def get_board_name(host_url, username, token, board_id):
    try:
        url = "https://{0}/rest/agile/1.0/board?startAt=0".format(host_url)
        headers = {"Accept": "application/json", "Content-Type": "application/json"}
        response = requests.get(url, headers=headers, auth=(username, token))
        logger.info(response.status_code)
        if response.status_code == 200:
            data = response.json()
            num_of_itr = round(data["total"] / 50 + 1)
            boardiddict = {}
            for i in range(0, num_of_itr + 1):
                startAt = i * 50
                url = "https://{1}/rest/agile/1.0/board?startAt={0}".format(
                    startAt, host_url
                )
                response = requests.get(url, headers=headers, auth=(username, token))
                data = response.json()
                for i in range(0, len(data["values"])):
                    boardiddict.update(
                        {data["values"][i]["id"]: data["values"][i]["name"]}
                    )
            board_name = boardiddict[board_id]
            return [board_name]
        elif response.status_code == 401:
            return [None, "Invalid username or password."]
        elif response.status_code == 403:
            return [
                None,
                "You do not have permission to access the requested resource.",
            ]
        elif response.status_code == 400:
            return [None, "Bad Request Payload Data"]
    except Exception as e:
        return [None, "Error while getting board id"]


def convert_seconds_to_time(seconds):
    # Calculate hours, minutes, and seconds
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    # Format the time string
    time_string = (
        f"{hours} hour{'s' if hours != 1 else ''}, "
        f"{minutes} minute{'s' if minutes != 1 else ''}"
    )
    return time_string


def time_conversion(timestamp_str):
    # Your original timestamp
    timestamp_str = timestamp_str

    # Parse the original timestamp
    # Note: '%z' directive works for the format '+HHMM' or '-HHMM' (without a colon)
    timestamp_dt = datetime.strptime(timestamp_str, "%Y-%m-%dT%H:%M:%S.%f%z")

    # Format the datetime object to just the date portion in "YYYY-MM-DD" format
    formatted_date = timestamp_dt.strftime("%Y-%m-%d")
    return formatted_date


def compare_date_with_today(input_date_str):
    """Function to compare a given date with today's date"""
    # Parse the input date string into a datetime object
    input_date = datetime.strptime(input_date_str, "%Y-%m-%d")
    # Get today's date
    today = datetime.today()
    # Compare the two dates
    if input_date.date() < today.date():
        return False
    else:
        return True
    
def convert_seconds(seconds):
    # Convert seconds to hours
    hours = seconds / 3600

    # Convert hours to days, where 8 hours equal 1 day
    days = hours // 8
    remaining_hours = hours % 8

    return f"{int(days)}d,{int(remaining_hours)}h"
